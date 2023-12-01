import tkinter as tk
from tkinter import filedialog
import openpyxl

class ExcelRow:
    def __init__(self, **kwargs):
        for column, value in kwargs.items():
            setattr(self, str(column), value)

def dot_counter(input_string):
    return input_string.count('.')

class module_class:
    def __init__(self, name):
        self.name = name
        self.tables = []


class parameter_class:
    def __init__(self, display_name, parameter_type,
                 mandatory_range, gui_value, actual_value, enumeration_type, default_value,
                 initial_value, description, version, access_mode, parameter, service_affect,
                 rules, notlar):

        self.display_name = display_name
        self.parameter_type = parameter_type
        self.mandatory_range = mandatory_range
        self.gui_value = gui_value
        self.actual_value = actual_value
        self.enumeration_type = enumeration_type
        self.default_value = default_value
        self.initial_value = initial_value
        self.description = description
        self.version = version
        self.access_mode = access_mode
        self.parameter = parameter
        self.service_affect = service_affect
        self.rules = rules
        self.notlar = notlar

def create_parameter_instance(parameter_values):
    return parameter_class(*parameter_values)

class table_class:
    def __init__(self, name):
        self.name = name
        self.inside_table = []
        self.parameter = []



def dot_check(my_string):
    dot_index = my_string.find(".")
    return dot_index

module_names =[]
modules = []
table_names =[]
tables = []
def read_excel_file(file_path):
    try:

        workbook = openpyxl.load_workbook(file_path, data_only=True)

        excel_data = {}

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            sheet_data = []
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
                i = 0
                for cell in row:
                    if i == 0 or i == 1 or i == 2:   #for module and table columns
                        if i == 0:

                            if cell == None:
                                i = i + 1
                                continue
                            else:
                                modules.append(module_class(name = cell))
                                module_names.append(cell)

                        if i == 1:
                            if cell == None:
                                i = i + 1
                                continue
                            else:
                                if dot_check(cell) == -1:       #deep 1
                                    double_deep = 0
                                    #(modules[-1].tables).append(table_class(name= cell))

                                else:                           #deep 2
                                    double_deep = 1
                                    index = dot_check(cell)
                                   # (modules[-1].tables[-1].inside_table).append(table_class(name=cell[index + 1:]))


                                 #   if dot_check(cell[index + 1:]) == -1:
                                  #      double_deep = 0
                                   #     (modules[-1].tables[-1].inside_table).append(table_class(name=cell[index + 1:]))
                                    #    if (modules[-1].tables[-1].name == cell[:index]):
                                     #       print("control_point")
                                      #      item = create_parameter_instance(row[3:18])
                                      #   (modules[-1].tables[-1].inside_table[-1].parameter).append(item)

                                    #else: #deep2
                                     #   double_deep = 1
                                      #  new_index = dot_check(cell[index + 1:])
                                       # if dot_check(cell[index + 1:]) == -1:
                                       #     (modules[-1].tables[-1].inside_table[-1].inside_table).append(table_class(name=cell[new_index + 1:]))
                                        #    if ( modules[-1].tables[-1].inside_table[-1].name == cell[index + 1: new_index] ):
                                         #       (modules[-1].tables[-1].inside_table[-1].parameter).append(create_parameter_instance(row[3:18]))

                        if i == 2:
                            if (cell != None):
                                depht = dot_counter(cell)
                                index = dot_check(cell)
                            for d in range(5 + 1):
                                if double_deep == 1:
                                    if d == 5:
                                        f= 4
                                        #(modules[-1].tables[-1].inside_table[-1].inside_table[-1].parameter).append(create_parameter_instance(row[3:18]))
                                    else:
                                        f=3
                                       # (modules[-1].tables[-1].inside_table[-1].inside_table).append(table_class(name=cell[:index]))



                                else:
                                    if d == 5:
                                        g=0
                                        #(modules[-1].tables[-1].inside_table[-1].parameter).append(create_parameter_instance(row[3:18]))
                                    else:
                                        f=4
                                        #(modules[-1].tables[-1].inside_table).append(table_class(name=cell[:index]))



                        if i == 2:
                            f = 4
                    i = i + 1





            v = 0
            for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column,values_only=True):
                for cell in col:
                    if cell:
                        if v == 0:
                            modules.append(module_class(cell))
                            module_names.append(cell)

                        if v == 1:
                            table_names.append(cell)

                    else:
                        continue

                v = v + 1



            excel_data[sheet_name] = sheet_data


        workbook.close()
        #print("*********")
        #print(module_names)
        #print("^^^^^^^^")
        #print(table_names)

        return excel_data

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


import xml.etree.ElementTree as ET
from xml.dom import minidom










def open_excel_file():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    excel_data = read_excel_file(file_path)
    #generateXSD("xsddd_example.xsd")

    # print_excel_data(excel_data)

if __name__ == "__main__":
    open_excel_file()
