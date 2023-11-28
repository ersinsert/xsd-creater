import tkinter as tk
from tkinter import filedialog
import openpyxl
import warnings
warnings.simplefilter("ignore", UserWarning)




def read_excel_file(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)

        # Print the names of all sheets in the workbook
        sheet_names = workbook.sheetnames
        print("Sheet Names:", sheet_names)

        # Assuming you want to print the values in each cell of the first sheet
        sheet = workbook.active

        if sheet is not None:
            for row in sheet.iter_rows(values_only=True):
                print(row)
        else:
            print("No active sheet found in the workbook.")

        workbook.close()

    except Exception as e:
        print(f"An error occurred: {e}")

def open_excel_file():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    if file_path:
        read_excel_file(file_path)
    else:
        print("No file selected.")

if __name__ == "__main__":
    open_excel_file()
