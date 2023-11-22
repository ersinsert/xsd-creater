import tkinter as tk
from tkinter import filedialog
import openpyxl

class ExcelRow:
    def __init__(self, **kwargs):
        for column, value in kwargs.items():
            setattr(self, str(column), value)

class Parameter:
    def __init__(self, module, table_name, parameter_name, display_name, parameter_type,
                 mandatory_range, gui_value, actual_value, enumeration_type, default_value,
                 initial_value, description, version, access_mode, parameter, service_affect,
                 rules, notlar):
        self.module = module
        self.table_name = table_name
        self.parameter_name = parameter_name
        self.display_name = display_name
        self.parameter_type = parameter_type
        self.mandatory_range = mandatory_range
        self.gui_value = gui_value
        self.actual_value = actual_value
        self.enumeration_type = enumeration_type
        self.default_value = default_value
        self.initial_value = initial_value
        self.description = description
        self.version = version
        self.access_mode = access_mode
        self.parameter = parameter
        self.service_affect = service_affect
        self.rules = rules
        self.notlar = notlar




def read_excel_file(file_path):
    try:
        # Open the selected Excel file
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        excel_data = {}  # Dictionary to store sheets

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Get the column names from the first row
            columns = [cell.value for cell in sheet[1]]

            # Create a list to store the data from the current sheet
            sheet1 = []

            for row_data in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
                # Create an ExcelRow instance for each row
                print(row_data)
                sheet1.append(row_data[0], row_data[1], row_data[2], row_data[3], row_data[4], row_data[5], row_data[6],
                              row_data[7], row_data[8], row_data[9], row_data[10], row_data[11], row_data[12],
                              row_data[13], row_data[14], row_data[15], row_data[16])





        # Close the Excel file
        workbook.close()

        return excel_data
    except Exception as e:
        print(f"An error occurred: {e}")
        return None

def open_excel_file():
    root = tk.Tk()
    root.withdraw()

    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    if file_path:
        excel_data = read_excel_file(file_path)
        if excel_data is not None:
            for sheet_name, sheet_data in excel_data.items():
                print(f"Sheet: {sheet_name}")
                for row in sheet_data:
                    # Access data using attribute names (e.g., row.modules)
                    print(vars(row))
                print()
    else:
        print("No file selected.")

if __name__ == "__main__":
    open_excel_file()
